from openpyxl import load_workbook
import os
import json
import copy
from bintang.table import Table, Table_Path
from bintang.table import _match_primitive, _match_caseless, _match_caseless_unicode
from bintang import iterdict
from pathlib import Path
from bintang.log import log
from rapidfuzz import fuzz , process, utils

# import logging

# log = logging.getLogger(__name__)
# FORMAT = "[%(filename)s:%(lineno)s - %(funcName)10s() ] %(message)s"
# logging.basicConfig(format=FORMAT)
# log.setLevel(logging.DEBUG)
# class TableNotFoundError(Exception):
#     def __init__(self,tablename):
#         tablenames = self.get_tables()
#         suggest = process.extract(tablename, tablenames, limit=2)
#         print('hello',suggest)
#         self.message = "Cannot find table '{}'.".format(tablename)
#         super().__init__(self.message)


class Bintang():
    def __init__(self, name=None, backend=None):
        self.name = name
        self.parent = 'dad'
        self.__tables = {} # this must be a dict of id:table object
        self.__last_assigned_tableid= -1 #
        self.__be = None # will be deprecated
        if backend is not None: # will be deprecated
            from bintang.besqlite import Besqlite # will be deprecated
            self.__be = Besqlite(self.name) # will be deprecated
 

    def __getitem__(self, tablename: str) -> Table: # subscriptable version of self.get_table()
        tableid = self.get_tableid(tablename)
        if tableid is None:
            tablenames = self.get_tables()
            extracted = process.extract(tablename, tablenames, limit=2, processor=utils.default_process)
            fuzzies = [repr(x[0]) for x in extracted if x[1] > 75]
            if len(fuzzies) > 0:
                raise ValueError ('could not find table {}. Did you mean {}?'.format(repr(tablename),' or '.join(fuzzies)))
            else:
                raise ValueError ('could not find table {}.'.format(repr(tablename)))
        else:
            return self.__tables[tableid]       

    def __repr__(self):
        rb = {}
        rb['name'] = self.name
        table = []
        for tablename in self.get_tables():
            table.append(tablename)
        rb['tables'] = table
        return json.dumps(rb, indent=2)

    

    def __del__(self):
        if self.__be is not None:
            self.__be.conn.close() #win need this otherwise a PermissionError: [WinError 32] ...
            #os.remove(self.__be.dbpath)
        

    def copy_db(self, dest=None):
        import shutil
        if dest == None:
            dest = os.getcwd()
        shutil.copy(self.__be.dbpath, dest)      


    def create_table(self, name: str, 
                     columns: list=None,
                     conn: str=None) -> Table:
        """ create a table under bintang object
        name: Name of the table
        columns: List of columns (optional)"""
        tobj = Table(name, bing=self, conn=conn) # create a tobj object
        self.add_table(tobj)
        if self.__be is not None:   # if is_persistent is True then update the tobj attributes and pass the connection
            tobj._Table__be = self.__be           
            tobj._Table__be.add_table(self.get_tableid(name), name)
        if columns is not None:
            for column in columns: # add column
                tobj.add_column(column)
        return tobj        

    def drop_table(self, tablename):
        tableid = self.get_tableid(tablename)
        if tableid is None:
            tablenames = self.get_tables()
            extracted = process.extract(tablename, tablenames, limit=2, processor=utils.default_process)
            fuzzies = [repr(x[0]) for x in extracted if x[1] > 75]
            raise ValueError ('could not find table {}. Did you mean {}?'.format(repr(tablename),' or '.join(fuzzies)))
        else:
            del self.__tables[tableid]


    def create_path_table(self, name, columns=None):
        tobj = Table_Path(name, bing=self) # create a table object
        self.add_table(tobj)
        if self.__be is not None:   # if is_persistent is True then update the tobj attributes and pass the connection
            tobj._Table__be = self.__be           
            tobj._Table__be.add_table(self.get_tableid(name), name)
        if columns is not None:
            for column in columns: # add column
                tobj.add_column(column)            
        
        
    def get_tableid(self, tablename):
        for id, table in self.__tables.items():
            if tablename.upper() == table.name.upper(): # compare uppercased to ensure case insensitivity
                return id
        return None


    def get_tables(self):
        tables = []
        for table in self.__tables.values():
            tablename = table.name
            tables.append(tablename)    
        return tables


    def get_tablepaths(self):
        pathnames = []
        for table in self.__tables.values():
            pathname = table.path
            pathnames.append(pathname)    
        return pathnames    

    
    def get_columns(self, tablename):
        return self.get_table(tablename).get_columns()


    # def get_columnids(self, tablename, columns):
    #     tableid = self.get_tableid(tablename)
    #     return self.__tables[tableid].get_columnids
    

    def add_table(self,table: str): # this should be Table not str. change later after tests
        tableid = self.get_tableid(table.name)
        if tableid is None:
            tableid = self.__last_assigned_tableid + 1
            self.__tables[tableid] = table
            self.__last_assigned_tableid = self.__last_assigned_tableid + 1 # update it
        elif tableid is not None:
            raise ValueError('Table {} already exists.'.format(table.name))


    def rename_table(self, name, new_tablename):
        # check if name already exists
        if name in self.get_tables():
            for id, table in self.__tables.items():
                if name == table.name:
                    self.__tables[id].name = new_tablename
        else:
            raise ValueError('Tablename {} does not exist.'.format(name))

    def get_table(self, name):
        tableid = self.get_tableid(name)
        return self.__tables[tableid]


    def copy_table(self,source_tablename, destination_tablename):
        destination_table = copy.deepcopy(self.get_table(source_tablename))
        destination_table.name = destination_tablename
        self.add_table(destination_table)
        
    
    def drop_column(self,tablename,column):
        self.__tables[tablename].drop_column(column)


    def insert(self,tablename,columns,values):
        self.get_table(tablename).insert(columns,values)
        

    def _get_cells_bycolumns(self,tablename,row,columns):
        _cells = {} # to hold the results cells
        for name in columns:
            columnid = self.__tables[tablename].get_columnid(name)
            # only assign if columnid exists to avoid a KeyError
            if columnid in row.cells:
                _cells[columnid] = row.cells[columnid]
        return _cells

    def delete_row(self, tablename, index):
        self.get_table(tablename).delete_row(index)

    
    def delete_rows(self, tablename, indexes):
        self.get_table(tablename).delete_rows(indexes)


    def print(self):
        tobj = Table('Columns Info')
        row_dict = {}
        for tab in self.get_tables():
            self[tab].set_data_props() # it sets 
            row_dict['table'] = tab
            for col in self[tab].get_columns():
                cobj = self[tab].get_column_object(col)
                #row_dict['id'] = cobj.id
                row_dict['column'] = cobj.name
                row_dict['ordinal_position'] = cobj.ordinal_position
                row_dict['data_type'] = cobj.data_type
                row_dict['column_size'] = cobj.column_size
                row_dict['decimal_digits'] = cobj.decimal_digits
                row_dict['data_props'] = cobj.data_props
                tobj.insert(row_dict)
        tobj.print()


    def raise_valueerror_tablename(self,tablename):
        tableid = self.get_tableid(tablename)
        if tableid is None:
            #tablenames = self.get_tables()
            # suggest = process.extract(tablename, tablenames, limit=2)
            # print('hello',suggest)
            # print('yes')
            # quit()
            raise ValueError('Tablename {} does not exist.'.format(tablename))


    def iterrows(self, table, columns=None, row_type = 'dict', rowid=False):
        """get the table form the collection
        then yield idx and row from table's iterrows()
        """
        
        #self.raise_valueerror_tablename(tablename)
        if isinstance(table, Table):
            for idx, row in table.iterrows(columns, row_type=row_type, rowid=rowid):
                yield idx, row
        else:
            for idx, row in self.get_table(table).iterrows(columns, row_type=row_type, rowid=rowid):
                yield idx, row
                

    def dev_read_db(self, connstr, tablename, columns = None):
        self.create_table(tablename)
        self[tablename].connstr = connstr
        import pyodbc
        conn = pyodbc.connect(connstr)
        cursor = conn.cursor()
        for row in cursor.tables():
            log.debug(row.table_name)
        if cursor.tables(table=tablename.lower()).fetchone():
            log.debug('table {} exists!'.format(tablename))
        else:
            log.debug('table {} not found'.format(tablename))
            # create table for rows
            sql_ = 'CREATE TABLE {} (idx serial NOT NULL PRIMARY KEY, id serial NOT NULL, cells json)'.format(tablename.lower())
            cursor.execute(sql_)
            cursor.commit()
            # create table for columns
            sql_ = 'CREATE TABLE {} (id serial NOT NULL PRIMARY KEY, name varchar(150))'.format(tablename.lower()+'_columns')
            cursor.execute(sql_)
            cursor.commit()
            # add columns if passed
            if columns is not None:
                if len(columns) > 0:
                    for col in columns:
                        sql_ = "INSERT INTO {} (name) VALUES ('{}');".format(tablename.lower()+'_columns', col)
                        log.debug(sql_)
                        cursor.execute(sql_)
                        cursor.commit()
        cursor.close()
        conn.close()    


    def _dev_read_sqlite(self, tablename, columns = None):
        self.create_table(tablename)
        import sqlite3
        conn = sqlite3.connect('bintang.db')
        cur = conn.cursor()
        sql_ = "SELECT EXISTS (SELECT 1 FROM sqlite_master WHERE type='table' AND name='{}')".format(tablename)
        ret = cur.execute(sql_)
        for row in cur.execute('SELECT * FROM sqlite_master'):
            log.debug(row)


        conn.close()    
        quit()
        if True:
            log.debug('table {} exists!'.format(tablename))
        else:
            log.debug('table {} not found'.format(tablename))
            # create table for rows
            sql_ = 'CREATE TABLE {} (id serial NOT NULL, cells json)'.format(tablename.lower())
            cursor.execute(sql_)
            # create table for columns
            sql_ = 'CREATE TABLE {} (id serial NOT NULL PRIMARY KEY, name varchar(150))'.format(tablename.lower()+'_columns')
            cursor.execute(sql_)
            # add columns if passed
            if columns is not None:
                if len(columns) > 0:
                    for col in columns:
                        sql_ = "INSERT INTO {} (name) VALUES ('{}');".format(tablename.lower()+'_columns', col)
                        log.debug(sql_)
                        cursor.execute(sql_)
        cursor.close()
        conn.close() 


    def _read_excel_OLD(self, path, sheetname, table=None):
        table_ = sheetname
        if table is not None:
            table_ = table
        if table_ not in self.get_tables():
            self.create_table(table_)        
        if self.__be is not None:
            self.__be.create_table(table_)
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheetname]
        columns = []
        Nonecolumn_cnt = 0
        for rownum, row_cells in enumerate(ws.iter_rows(),start=1):
            values = [] # hold column value for each row
            if rownum == 1:
                for cell in row_cells:
                    if cell.value is None:
                        columname = 'noname' + str(Nonecolumn_cnt)
                        Nonecolumn_cnt += 1
                        columns.append(columname)
                    else:
                        columns.append(cell.value)
                if Nonecolumn_cnt > 0:
                    log.warning('Warning! Noname column detected!')          
            
            if rownum > 1:
                for cell in row_cells:
                    values.append(cell.value)
                # if rownum == 370:
                #     log.debug(f'{values} at rownum 370.')
                #     log.debug(any(values))
                if any(values):
                    self.get_table(table_).insert(values, columns)
        if self.__be is not None:
            self.get_table(table_).add_row_into_be()


    def read_excel(self, path, sheetnames=None):
        wb = load_workbook(path, read_only=True, data_only=True)
        # validate sheetnames
        sheetnames_lced = {x.lower(): x  for x in wb.sheetnames}
        if sheetnames is not None: # user specify sheets
            for sheetname in sheetnames:
                if sheetname.lower() not in sheetnames_lced:
                    extracted = process.extract(sheetname, sheetnames_lced.values(), limit=2, processor=utils.default_process)
                    fuzzies = [repr(x[0]) for x in extracted if x[1] > 75]
                    raise ValueError ('could not find sheetname {}. Did you mean {}?'.format(repr(sheetname),' or '.join(fuzzies)))
                self.create_table(sheetname)
                self[sheetname].read_excel(path, sheetname)
        else: # exctract all sheets and create all tables
            for ws in wb:
                sheetname = ws.title
                self.create_table(sheetname)
                self[sheetname].read_excel(path, sheetname)


    def read_dict(self, dict_obj, tablepaths=None):
        debug = False
        for tprow in iterdict.iterdict(dict_obj, tablepaths):
            # if debug:
            #     print("\n---------------------in bintang---------------------")
            #     print(row)
            
            for k,v in tprow.cells.items():
                # print(k,'->',v)
                
                # create a table if not created yet
                if tprow.tablepath not in self.get_tables():
                    self.create_path_table(tprow.tablepath)
                
                # upsert this row
                self.get_table(tprow.tablepath).upsert_table_path_row(tprow)
            # if debug:
            #     print("---------------------out bintang--------------------")


    def read_json(self, json_str, tablepaths=None):
        dict_obj = json.loads(json_str) # parse json_str
        self.read_dict(dict_obj, tablepaths) 


    def set_child_tables(self):
        for tab1 in self.get_tablepaths():
            print('table name',tab1)
            # print(self[tab1].get_path_aslist())
            tab1_pathlist = self[tab1].get_path_aslist()
            for tab2 in self.get_tablepaths():
                tab2_pathlist = self[tab2].get_path_aslist()
                if len(tab2_pathlist) - len(tab1_pathlist) == 1:
                    # this is a possible child, not grand child
                    matches = 0
                    for i in range(len(tab1_pathlist)):
                        if tab1_pathlist[i] == tab2_pathlist[i]:
                            matches += 1
                    if matches == len(tab1_pathlist):
                        # this is a child
                        self[tab1].children.append(tab2)      


    def read_sql(self, conn, tablename, sql_str, params=None ):
        self.create_table(tablename)
        cursor = conn.cursor()
        if params is not None:
            cursor.execute(sql_str, params)
        else:
            cursor.execute(sql_str)
        columns = [col[0] for col in cursor.description]
        for row in cursor.fetchall():
            self.insert(tablename, row, columns)
             

    def VOID_get_row_asdict(self, tablename, idx, columns=None):
        return self.get_table(tablename).get_row_asdict(idx, columns=columns)


    def _add_lcell(self, lidx, ltable, outrow, out_table, out_lcolumns, rowid):
        for k, v in self[ltable].get_row_asdict(lidx,rowid=rowid).items():
            if out_lcolumns is None:
                # incude all columns
                cell = self[out_table].make_cell(k,v)
                outrow.add_cell(cell)
            if out_lcolumns is not None:
                # include only the passed columns
                if k in out_lcolumns:
                    cell = self[out_table].make_cell(k,v)
                    outrow.add_cell(cell)
        return outrow            


    def _add_rcell(self, ridx, rtable, outrow, out_table, out_rcolumns, rcol_resolved, rowid):
        for k, v in rtable.get_row_asdict(ridx,rowid=rowid).items():
            if out_rcolumns is None:
                cell = self[out_table].make_cell(rcol_resolved[k],v)
                outrow.add_cell(cell)
            if out_rcolumns is not None:
                # include only the passed columns
                if k in out_rcolumns:
                    cell = self[out_table].make_cell(rcol_resolved[k],v)
                    outrow.add_cell(cell)
        return outrow


    def _resolve_join_columns (self,ltable, rtable_obj, rowid=False):
        # resolve any conflict column by prefixing its tablename
        # notes: conflict column occurs when the same column being used in two joining table.
        rcol_resolved = {}
        if rowid:
            rcol_resolved['rowid_'] = ltable + '_' + 'rowid_'
        lcolumns = self[ltable].get_columns()
        for col in rtable_obj.get_columns():
            if col in lcolumns:
                rcol_resolved[col] = rtable_obj.name + '_' + col
            else:
                rcol_resolved[col] = col
        return rcol_resolved  



    # def _DEPRECATED_innerjoin(self,ltable, lkeys
    #             ,rtable, rkeys
    #             ,into
    #             ,out_lcolumns=None
    #             ,out_rcolumns=None
    #             ,rowid=False):

    #     # validate input eg. column etc
    #     lkeys = self[ltable].validate_columns(lkeys)
    #     rkeys = self[rtable].validate_columns(rkeys)
    #     if out_lcolumns is not None:
    #         out_lcolumns = self[ltable].validate_columns(out_lcolumns)
    #     if out_rcolumns is not None:
    #         out_rcolumns = self[rtable].validate_columns(out_rcolumns)

    #     # resolve columns conflicts
    #     rcol_resolved = self._resolve_join_columns(ltable, rtable, rowid)
    #     # create an output table
    #     out_table = into
    #     self.create_table(out_table)
    #     out_tobj = self.get_table(out_table)

    #     # for debuging create merged table to store the matching rowids
    #     #merged = self.create_table("merged",["lrowid","rrowid"])

    #     numof_keys = len(lkeys)
    #     # loop left table
    #     for lidx, lrow in self.iterrows(ltable, columns=lkeys, rowid=rowid):
    #         # loop right table
    #         for ridx, rrow in self.iterrows(rtable, columns=rkeys, rowid=rowid):
    #             matches = 0 # store matches for each rrow
    #             # compare value for any matching keys, if TRUE then increment matches
    #             for i in range(numof_keys):
    #                 if lrow[lkeys[i]] == rrow[rkeys[i]]:
    #                     matches += 1 # incremented!
    #             if matches == numof_keys: # if fully matched, create the row & add into the output table
    #                 #debug merged.insert(["lrowid","rrowid"], [lrow["_rowid"], rrow["_rowid"]])
    #                 #and add the row to output table
    #                 outrow = out_tobj.make_row()
    #                 # add cells from left table
    #                 outrow = self._add_lcell(lidx, ltable, outrow, out_table, out_lcolumns, rowid)
    #                 # add cells from right table
    #                 outrow = self._add_rcell(ridx, rtable, outrow, out_table, out_rcolumns, rcol_resolved, rowid)   
    #                 out_tobj.add_row(outrow)
    #     #debug merged.print() 
    #     return out_tobj


    def innerjoin(self
                ,ltable: str #, lkeys
                ,rtable: str | Table #, rkeys
                ,on: list[tuple] # list of lkey & r key tuple
                ,into: str
                ,out_lcolumns: list=None
                ,out_rcolumns: list=None
                ,rowid=False) -> Table:
        
        rtable_obj = None
        if isinstance(rtable,Table):
            rtable_obj = rtable
            rtable = rtable.name
        else:
            rtable_obj = self[rtable]
        # validate input eg. column etc
        lkeys = [x[0] for x in on] # generate lkeys from on (1st sequence)
        rkeys = [x[1] for x in on] # generate rkeys from on (2nd sequence)
        lkeys = self[ltable].validate_columns(lkeys)
        #rkeys = self[rtable].validate_columns(rkeys)
        rkeys = rtable_obj.validate_columns(rkeys)
        if out_lcolumns is not None:
            out_lcolumns = self[ltable].validate_columns(out_lcolumns)
        if out_rcolumns is not None:
            # out_rcolumns = self[rtable].validate_columns(out_rcolumns)
            out_rcolumns = rtable_obj.validate_columns(out_rcolumns)

        # resolve columns conflicts
        rcol_resolved = self._resolve_join_columns(ltable, rtable_obj, rowid)
        # create an output table
        out_table = into
        self.create_table(out_table)
        out_tobj = self.get_table(out_table)

        # for debuging create merged table to store the matching rowids
        #merged = self.create_table("merged",["lrowid","rrowid"])

        numof_keys = len(on) #(lkeys)
        # loop left table
        for lidx, lrow in self.iterrows(ltable, columns=lkeys, rowid=rowid):
            # loop right table
            for ridx, rrow in self.iterrows(rtable_obj, columns=rkeys, rowid=rowid):
                matches = 0 # store matches for each rrow
                # compare value for any matching keys, if TRUE then increment matches
                for i in range(numof_keys): 
                    if _match_caseless_unicode(lrow[lkeys[i]], rrow[rkeys[i]]):
                        matches += 1 # incremented!
                if matches == numof_keys: # if fully matched, create the row & add into the output table
                    #debug merged.insert(["lrowid","rrowid"], [lrow["_rowid"], rrow["_rowid"]])
                    #and add the row to output table
                    outrow = out_tobj.make_row()
                    # add cells from left table
                    outrow = self._add_lcell(lidx, ltable, outrow, out_table, out_lcolumns, rowid)
                    # add cells from right table
                    outrow = self._add_rcell(ridx, rtable_obj, outrow, out_table, out_rcolumns, rcol_resolved, rowid)   
                    out_tobj.add_row(outrow)
        #debug merged.print() 
        return out_tobj
                

    def leftjoin(self,ltable, rtable, lkeys, rkeys
                ,out_lcolumns=None
                ,out_rcolumns=None
                ,rowid=False):

        # validate input eg. column etc
        lkeys = self[ltable].validate_columns(lkeys)
        rkeys = self[rtable].validate_columns(rkeys)
        if out_lcolumns is not None:
            out_lcolumns = self[ltable].validate_columns(out_lcolumns)
        if out_rcolumns is not None:
            out_rcolumns = self[rtable].validate_columns(out_rcolumns)
        

        # create an output table 
        out_table = ltable + rtable
        self.create_table(out_table)
        out_tobj = self.get_table(out_table)

        # for debuging create merged table to store the matching rowids
        #merged = self.create_table("merged",["lrowid","rrowid"])

        numof_keys = len(lkeys)
        for lidx, lrow in self.iterrows(ltable, columns=lkeys, rowid=rowid):
            outrow = out_tobj.make_row()
            # add cells for ltable
            outrow = self._add_lcell(lidx, ltable, outrow, out_table, out_lcolumns, rowid)
            for ridx, rrow in self.iterrows(rtable, columns=rkeys, rowid=rowid):
                matches = 0 # store matches for each rrow
                # evaluate any matching keys, if so increment matches
                for i in range(numof_keys):
                    if _match_caseless_unicode(lrow[lkeys[i]] == rrow[rkeys[i]]):
                        matches += 1 # increment
                if matches == numof_keys: # if fully matched, create the row & add into the output table
                    #debug merged.insert(["lrowid","rrowid"], [lrow["_rowid"], rrow["_rowid"]])
                    # add cells for rtable
                    outrow = self._add_rcell(ridx, rtable, outrow, out_table, out_rcolumns, ltable, rowid)   
            out_tobj.add_row(outrow)             
        #debug merged.print() 
        return out_tobj
    